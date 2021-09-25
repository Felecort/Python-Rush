import openpyxl

xls_parth = "D:\\Projects\\PythonProjects\\Learn\\words.xlsx"
txt_parth_default = "D:\\Projects\\PythonProjects\\Learn\\words.txt"
wookbook = openpyxl.load_workbook(xls_parth)
worksheet = wookbook.active


def write_in_file(text):
    with open(txt_parth_default, "w", encoding="UTF-8") as file:
        file.write(text)
        

def pars_default_words():
    rows = worksheet.max_row
    i = 3
    text = ""
    while True:
        row_arr = []
        for j in range(2, 4):
            cell = worksheet.cell(row=i, column=j).value
            if not cell:
                write_in_file(text)
                return
            row_arr.append(cell)
        text += f'\"{row_arr[0]}\";\"{row_arr[1]}\"\n'
        row_arr.clear()
        i += 1


pars_default_words()
